#!/usr/bin/env python3
"""
品牌档案数据源.xlsx · sheet「类目数据源」→ brand_category_monthly.csv

口径（二级品类 × 品牌 × 月）：
- 品牌下类目占比 = 该二级品类四渠道销售额合计 / 品牌四渠道销售额合计（各类目加总 100%）
- JD品类占比 = 该二级品类 JD 销售额 / 品牌 JD 销售额合计（各类目加总 100%）
- JD/TM/TB/DY = 该二级品类各渠道销售额 / 该品类四渠道销售额合计（各行加总 100%）
"""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("请先安装 openpyxl: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = Path("/Users/chenjiaxuan12/Desktop/品牌档案数据源.xlsx")
LOCAL_XLSX = ROOT / "品牌档案数据源.xlsx"
OUT_CSV = ROOT / "brand_category_monthly.csv"

SHEET = "类目数据源"
CHANNELS = {
    "JD": "jd",
    "A-TM(大贸+国内现货)": "tm",
    "A-TB(非国际)": "tb",
    "N-DY(非国际)": "dy",
}
BRAND_KEYS = {
    "jomoo", "arrow", "hegii", "submarine", "micoe",
    "nippon", "skshu", "dulux", "wacker", "yuhong", "carpoly",
}
MASTER_JSON = ROOT.parent / "brands_master.json"


def load_placeholder_map() -> dict[str, str]:
    if not MASTER_JSON.exists():
        return {}
    data = json.loads(MASTER_JSON.read_text(encoding="utf-8"))
    return {k: v for k, v in (data.get("placeholder_to_name_key") or {}).items()}


PLACEHOLDER_MAP = load_placeholder_map()


def resolve_brand_key(raw_key: str | None, name_key_col: str | None = None) -> str | None:
    if name_key_col:
        nk = name_key_col.strip().lower()
        if nk in BRAND_KEYS:
            return nk
    if not raw_key:
        return None
    key = str(raw_key).strip().lower()
    key = PLACEHOLDER_MAP.get(key, key)
    return key if key in BRAND_KEYS else None

CSV_FIELDS = ["name_key", "period_value", "category_distribution", "category_share"]


def parse_period(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    s = str(v).strip()
    m = re.match(r"(\d{4})年(\d{1,2})月", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    if len(s) >= 7 and s[4] == "-":
        return s[:7]
    return None


def pct(part: float, whole: float) -> float:
    if not whole:
        return 0.0
    return round(part / whole * 100, 1)


def build_category_payload(sales: dict[str, dict[str, float]]) -> tuple[list, list]:
    """sales: {二级品类: {jd,tm,tb,dy}}"""
    brand_total = sum(sum(ch.values()) for ch in sales.values())
    jd_total = sum(ch.get("jd", 0.0) for ch in sales.values())

    dist_rows = []
    share_rows = []
    for cat, ch in sales.items():
        cat_total = sum(ch.get(k, 0.0) for k in CHANNELS.values())
        if cat_total <= 0:
            continue
        jd = ch.get("jd", 0.0)
        tm = ch.get("tm", 0.0)
        tb = ch.get("tb", 0.0)
        dy = ch.get("dy", 0.0)
        dist_rows.append(
            {"name": cat, "share": pct(cat_total, brand_total), "sales": round(cat_total, 2)}
        )
        share_rows.append(
            {
                "name": cat,
                "jd_brand_share": pct(jd, jd_total),
                "jd": pct(jd, cat_total),
                "tm": pct(tm, cat_total),
                "tb": pct(tb, cat_total),
                "dy": pct(dy, cat_total),
            }
        )

    dist_rows.sort(key=lambda x: (-x["share"], x["name"]))
    order = {r["name"]: i for i, r in enumerate(dist_rows)}
    share_rows.sort(key=lambda x: order.get(x["name"], 999))
    return dist_rows, share_rows


def transform(xlsx_path: Path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"找不到 sheet「{SHEET}」")
    ws = wb[SHEET]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header) if h}

    raw: dict[tuple[str, str], dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(float))
    )

    for r in rows:
        nk_col = r[idx["name_key"]] if "name_key" in idx else None
        key = resolve_brand_key(r[idx["brand"]], str(nk_col) if nk_col else None)
        if not key:
            continue
        period = parse_period(r[idx["时间"]])
        if not period:
            continue
        ch_label = r[idx["渠道"]]
        if ch_label not in CHANNELS:
            continue
        cat = r[idx["二级品类"]]
        if not cat:
            continue
        amount = r[idx["销售额"]]
        if amount is None:
            continue
        raw[(key, period)][str(cat).strip()][CHANNELS[ch_label]] += float(amount)

    wb.close()

    out = []
    for (key, period) in sorted(raw.keys()):
        dist, share = build_category_payload(raw[(key, period)])
        if not dist:
            continue
        out.append(
            {
                "name_key": key,
                "period_value": period,
                "category_distribution": json.dumps(dist, ensure_ascii=False),
                "category_share": json.dumps(share, ensure_ascii=False),
            }
        )
    return out


def write_csv(rows, path: Path):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        w.writeheader()
        w.writerows(rows)


def main():
    xlsx = LOCAL_XLSX if LOCAL_XLSX.exists() else DEFAULT_XLSX
    if len(sys.argv) > 1:
        xlsx = Path(sys.argv[1])
    if not xlsx.exists():
        raise SystemExit(f"找不到数据源: {xlsx}")

    rows = transform(xlsx)
    write_csv(rows, OUT_CSV)
    print(f"已转换 {len(rows)} 行 → {OUT_CSV}")

    sample = [r for r in rows if r["name_key"] == "hegii" and r["period_value"] == "2026-05"]
    if sample:
        print("hegii 2026-05 类目数:", len(json.loads(sample[0]["category_distribution"])))


if __name__ == "__main__":
    main()
