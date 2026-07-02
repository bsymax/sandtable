#!/usr/bin/env python3
"""
基础建材 BI 导出 → brand_category_monthly.csv（三级类目）

口径与卫浴二级类目一致，仅颗粒度改为 **三级类目**：
- 品牌下类目占比 = 该三级品类四渠道销售额合计 / 品牌四渠道销售额合计（各类目加总 100%）
- JD品类占比 = 该三级品类 JD 销售额 / 品牌 JD 销售额合计（各类目加总 100%）
- JD/TM/TB/DY = 该三级品类各渠道销售额 / 该品类四渠道销售额合计（各行加总 100%）
"""

from __future__ import annotations

import csv
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("请先安装 openpyxl: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = Path("/Users/chenjiaxuan12/Desktop/基础建材1_.xlsx")
OUT_CSV = ROOT / "brand_category_monthly.csv"

JC_KEYS = {"nippon", "skshu", "dulux", "wacker", "yuhong", "carpoly"}
CHANNELS = {"jd": "jd", "tm": "tm", "tb": "tb", "dy": "dy"}

COL_DATE = 0
COL_CH = 1
COL_CAT3 = 7
COL_SALES = 11
COL_NK = 13


def pct(part: float, whole: float) -> float:
    if not whole:
        return 0.0
    return round(part / whole * 100, 1)


def build_category_payload(sales: dict[str, dict[str, float]]) -> tuple[list, list]:
    """sales: {三级品类: {jd,tm,tb,dy}}"""
    brand_total = sum(sum(ch.values()) for ch in sales.values())
    jd_total = sum(ch.get("jd", 0.0) for ch in sales.values())

    dist_rows = []
    share_rows = []
    for cat, ch in sales.items():
        cat_total = sum(ch.get(k, 0.0) for k in CHANNELS.values())
        if cat_total <= 0:
            continue
        jd = ch.get("jd", 0.0)
        tm = ch.get("tm", 0.0)
        tb = ch.get("tb", 0.0)
        dy = ch.get("dy", 0.0)
        dist_rows.append(
            {"name": cat, "share": pct(cat_total, brand_total), "sales": round(cat_total, 2)}
        )
        share_rows.append(
            {
                "name": cat,
                "jd_brand_share": pct(jd, jd_total),
                "jd": pct(jd, cat_total),
                "tm": pct(tm, cat_total),
                "tb": pct(tb, cat_total),
                "dy": pct(dy, cat_total),
            }
        )

    dist_rows.sort(key=lambda x: (-x["share"], x["name"]))
    order = {r["name"]: i for i, r in enumerate(dist_rows)}
    share_rows.sort(key=lambda x: order.get(x["name"], 999))
    return dist_rows, share_rows


def read_jc_sales(xlsx: Path):
    """返回 raw[(nk, period)][cat3][channel] = sales"""
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw: dict[tuple[str, str], dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(float))
    )
    for r in ws.iter_rows(min_row=2, values_only=True):
        nk = r[COL_NK]
        if not nk:
            continue
        nk = str(nk).strip().lower()
        if nk not in JC_KEYS:
            continue
        dt = r[COL_DATE]
        if not isinstance(dt, datetime):
            continue
        ch = str(r[COL_CH] or "").strip().lower()
        if ch not in CHANNELS:
            continue
        cat3 = str(r[COL_CAT3] or "").strip()
        if not cat3:
            continue
        period = dt.strftime("%Y-%m")
        sales = float(r[COL_SALES] or 0)
        raw[(nk, period)][cat3][CHANNELS[ch]] += sales
    wb.close()
    return raw


def build_rows(raw) -> list[dict]:
    rows = []
    for (nk, period) in sorted(raw.keys()):
        dist, share = build_category_payload(raw[(nk, period)])
        if not dist:
            continue
        rows.append(
            {
                "name_key": nk,
                "period_value": period,
                "category_distribution": json.dumps(dist, ensure_ascii=False),
                "category_share": json.dumps(share, ensure_ascii=False),
            }
        )
    return rows


def merge_csv(jc_rows: list[dict], out_path: Path):
    existing: list[dict] = []
    if out_path.exists():
        with out_path.open(newline="", encoding="utf-8") as f:
            existing = list(csv.DictReader(f))
    kept = [r for r in existing if r.get("name_key") not in JC_KEYS]
    merged = kept + jc_rows
    merged.sort(key=lambda r: (r["name_key"], r["period_value"]))
    fields = ["name_key", "period_value", "category_distribution", "category_share"]
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(merged)
    return len(jc_rows), len(kept), len(merged)


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"找不到: {xlsx}")
    raw = read_jc_sales(xlsx)
    jc_rows = build_rows(raw)
    n_jc, n_kept, n_total = merge_csv(jc_rows, OUT_CSV)
    sample = [r for r in jc_rows if r["name_key"] == "nippon" and r["period_value"] == "2026-05"]
    if sample:
        n = len(json.loads(sample[0]["category_distribution"]))
        print(f"nippon 2026-05 三级类目数: {n}")
    print(f"建材类目: {n_jc} 行 → {OUT_CSV}（保留卫浴 {n_kept} 行 · 合计 {n_total}）")


if __name__ == "__main__":
    main()
