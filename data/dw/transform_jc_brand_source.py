#!/usr/bin/env python3
"""
基础建材 BI 导出（类目长表）→ brand_metrics_monthly 行

源表：基础建材1_.xlsx（或传入路径）
粒度：标准品牌 name_key × 月 × 渠道（三级类目汇总）

口径（与佳璇确认 · 2026-07-01）：
- 成交 gmv / 销量 / SPU：均取 **JD 渠道** 汇总
- 「市占」jd/tm/tb/dy share = 该渠道销售额 ÷ **四渠道销售额合计** × 100
- 成交同比 / 销量同比 / 渠道增速：有 **去年同月** 数据则按 YoY 计算，否则留空
- jd_share_wow：JD 渠道占比 **环比** 变化（百分点 = 本月 jd_share − 上月 jd_share）
- sku_count：JD 渠道各行「去重计数_SPU」之和（类目间可能重复计数）

源表列（2026-07-01 更新）：第10列 **销量** · 第11列 **销售额** · 第13列 **name key**
  jd → JD · tm → 天猫 · tb → 淘宝 · dy → 抖音
"""

from __future__ import annotations

import csv
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("请先安装 openpyxl: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = Path("/Users/chenjiaxuan12/Desktop/基础建材1_.xlsx")
OUT_CSV = ROOT / "brand_metrics_monthly.csv"

JC_KEYS = {"nippon", "skshu", "dulux", "wacker", "yuhong", "carpoly"}

# 源表列索引（基础建材1_.xlsx）
COL_DATE = 0
COL_CH = 1
COL_VOL = 10
COL_SALES = 11
COL_SPU = 12
COL_NK = 13

CH_SRC = {"jd", "tm", "tb", "dy"}

CSV_FIELDS = [
    "name_key", "period_type", "period_value", "gmv", "gmv_yoy", "sales_volume",
    "sales_volume_yoy", "jd_share", "jd_share_wow", "tmall_share", "douyin_share",
    "taobao_share", "channel_growth_jd", "channel_growth_tmall", "channel_growth_douyin",
    "channel_growth_taobao", "sku_count",
]


def pct_ratio(numer: float, denom: float) -> float | None:
    if not denom:
        return None
    return round(numer / denom * 100, 2)


def yoy_pct(cur: float, prev: float) -> float | None:
    if prev is None or prev == 0:
        return None
    return round((cur - prev) / prev * 100, 2)


def wan(yuan: float) -> float:
    return round(yuan / 10000, 2)


def fmt(v):
    if v is None or v == "":
        return ""
    if isinstance(v, float):
        return f"{v:.2f}".rstrip("0").rstrip(".") if v % 1 else f"{int(v)}"
    return str(v)


def prev_month(period: str) -> str | None:
    y, m = map(int, period.split("-"))
    if m == 1:
        return f"{y - 1}-12"
    return f"{y}-{m - 1:02d}"


def same_month_last_year(period: str) -> str:
    y, m = period.split("-")
    return f"{int(y) - 1}-{m}"


def read_source(xlsx: Path) -> dict[tuple[str, str, str], dict]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    buckets: dict[tuple[str, str, str], dict] = defaultdict(
        lambda: {"sales": 0.0, "vol": 0.0, "spu": 0}
    )
    for r in ws.iter_rows(min_row=2, values_only=True):
        nk = r[COL_NK]
        if not nk:
            continue
        nk = str(nk).strip().lower()
        if nk not in JC_KEYS:
            continue
        dt = r[COL_DATE]
        if not isinstance(dt, datetime):
            continue
        period = dt.strftime("%Y-%m")
        ch = str(r[COL_CH] or "").strip().lower()
        if ch not in CH_SRC:
            continue
        sales = float(r[COL_SALES] or 0)
        vol = float(r[COL_VOL] or 0)
        spu = int(float(r[COL_SPU] or 0))
        b = buckets[(nk, period, ch)]
        b["sales"] += sales
        b["vol"] += vol
        b["spu"] += spu
    wb.close()
    return buckets


def build_brand_month_rows(buckets: dict) -> list[dict]:
    # 按品牌×月汇总四渠道
    by_bp: dict[tuple[str, str], dict[str, dict]] = defaultdict(dict)
    for (nk, period, ch), v in buckets.items():
        by_bp[(nk, period)][ch] = v

    # 先算各月 jd_share 供 wow
    jd_share_cache: dict[tuple[str, str], float | None] = {}
    for (nk, period), chs in sorted(by_bp.items()):
        total = sum(chs.get(c, {}).get("sales", 0) for c in CH_SRC)
        jd_sales = chs.get("jd", {}).get("sales", 0)
        jd_share_cache[(nk, period)] = pct_ratio(jd_sales, total)

    rows = []
    for (nk, period), chs in sorted(by_bp.items()):
        total = sum(chs.get(c, {}).get("sales", 0) for c in CH_SRC)
        jd = chs.get("jd", {"sales": 0, "vol": 0, "spu": 0})
        tm = chs.get("tm", {"sales": 0})
        tb = chs.get("tb", {"sales": 0})
        dy = chs.get("dy", {"sales": 0})

        ly = same_month_last_year(period)
        ly_chs = by_bp.get((nk, ly), {})

        jd_ly_sales = ly_chs.get("jd", {}).get("sales")
        jd_vol = jd.get("vol", 0)
        jd_ly_vol = ly_chs.get("jd", {}).get("vol")

        pm = prev_month(period)
        jd_share = jd_share_cache[(nk, period)]
        jd_share_prev = jd_share_cache.get((nk, pm)) if pm else None
        jd_share_wow = None
        if jd_share is not None and jd_share_prev is not None:
            jd_share_wow = round(jd_share - jd_share_prev, 2)

        def ch_yoy(ch: str) -> float | None:
            cur = chs.get(ch, {}).get("sales", 0)
            prev = ly_chs.get(ch, {}).get("sales")
            if prev is None:
                return None
            return yoy_pct(cur, prev)

        rows.append({
            "name_key": nk,
            "period_type": "monthly",
            "period_value": period,
            "gmv": wan(jd.get("sales", 0)),
            "gmv_yoy": yoy_pct(jd.get("sales", 0), jd_ly_sales) if jd_ly_sales is not None else None,
            "sales_volume": int(round(jd_vol)),
            "sales_volume_yoy": yoy_pct(jd_vol, jd_ly_vol) if jd_ly_vol is not None else None,
            "jd_share": jd_share,
            "jd_share_wow": jd_share_wow,
            "tmall_share": pct_ratio(tm.get("sales", 0), total),
            "douyin_share": pct_ratio(dy.get("sales", 0), total),
            "taobao_share": pct_ratio(tb.get("sales", 0), total),
            "channel_growth_jd": ch_yoy("jd"),
            "channel_growth_tmall": ch_yoy("tm"),
            "channel_growth_douyin": ch_yoy("dy"),
            "channel_growth_taobao": ch_yoy("tb"),
            "sku_count": jd.get("spu", 0) or None,
        })
    return rows


def merge_csv(jc_rows: list[dict], out_path: Path):
    existing: list[dict] = []
    if out_path.exists():
        with out_path.open(newline="", encoding="utf-8") as f:
            existing = list(csv.DictReader(f))
    kept = [r for r in existing if r.get("name_key") not in JC_KEYS]
    merged = kept + [{k: fmt(r.get(k)) for k in CSV_FIELDS} for r in jc_rows]
    merged.sort(key=lambda r: (r["name_key"], r["period_value"]))
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        w.writeheader()
        w.writerows(merged)
    return len(jc_rows), len(kept), len(merged)


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"找不到: {xlsx}")
    buckets = read_source(xlsx)
    jc_rows = build_brand_month_rows(buckets)
    n_jc, n_kept, n_total = merge_csv(jc_rows, OUT_CSV)
    periods = sorted({r["period_value"] for r in jc_rows})
    brands = sorted({r["name_key"] for r in jc_rows})
    print(f"基础建材: {n_jc} 行 · 品牌 {brands}")
    print(f"月份范围: {periods[0]} ~ {periods[-1]}（共 {len(periods)} 月）")
    print(f"合并 CSV: 保留原品牌 {n_kept} 行 + 建材 {n_jc} 行 = {n_total} 行 → {OUT_CSV}")
    # YoY 覆盖统计
    yoy_n = sum(1 for r in jc_rows if r.get("gmv_yoy") is not None)
    print(f"可计算成交同比的月份: {yoy_n}/{len(jc_rows)}（需源表含去年同月）")


if __name__ == "__main__":
    main()
