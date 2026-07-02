#!/usr/bin/env python3
"""
品牌档案数据源.xlsx → brand_metrics_monthly.csv

长表（品牌 × 月份 × 渠道）透视为数仓宽表（一行 = 品牌 × 月）。
渠道映射：
  JD → jd_share / jd_share_wow（市占及变化，供市占模块）
  JD 销售增速 → channel_growth_jd（与 TM/TB/DY 同为销售额同环比）
  A-TM(大贸+国内现货) → tmall_share / channel_growth_tmall
  A-TB(非国际) → taobao_share / channel_growth_taobao
  N-DY(非国际) → douyin_share / channel_growth_douyin

品牌级 KPI（成交/成交同比/销量/销量同比/动销 SPU/成交趋势）：均取 **JD 渠道**。
其他渠道仅用于渠道市占及 TM/DY/TB 渠道增速。
"""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("请先安装 openpyxl: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = Path("/Users/chenjiaxuan12/Desktop/品牌档案数据源.xlsx")
LOCAL_XLSX = ROOT / "品牌档案数据源.xlsx"
OUT_CSV = ROOT / "brand_metrics_monthly.csv"
OUT_BRANDS = ROOT / "brands_master.csv"

CHANNELS = {
    "JD": "jd",
    "A-TM(大贸+国内现货)": "tm",
    "A-TB(非国际)": "tb",
    "N-DY(非国际)": "dy",
}

CSV_FIELDS = [
    "name_key",
    "period_type",
    "period_value",
    "gmv",
    "gmv_yoy",
    "sales_volume",
    "sales_volume_yoy",
    "jd_share",
    "jd_share_wow",
    "tmall_share",
    "douyin_share",
    "taobao_share",
    "channel_growth_jd",
    "channel_growth_tmall",
    "channel_growth_douyin",
    "channel_growth_taobao",
    "sku_count",
]

BRAND_META = [
    ("jomoo", "九牧", "S", "周采销"),
    ("arrow", "箭牌", "A", "吴采销"),
    ("hegii", "恒洁", "A", "陈采销"),
    ("submarine", "潜水艇", "B", "李采销"),
    ("micoe", "四季沐歌", "B", "王采销"),
    ("nippon", "立邦", "B", "待定"),
    ("skshu", "三棵树", "B", "待定"),
    ("dulux", "多乐士", "B", "待定"),
    ("wacker", "瓦克", "B", "待定"),
    ("yuhong", "雨虹防水", "B", "待定"),
    ("carpoly", "嘉宝莉", "B", "待定"),
]

MASTER_JSON = ROOT.parent / "brands_master.json"


def load_placeholder_map() -> dict[str, str]:
    if not MASTER_JSON.exists():
        return {}
    data = json.loads(MASTER_JSON.read_text(encoding="utf-8"))
    return {k: v for k, v in (data.get("placeholder_to_name_key") or {}).items()}


PLACEHOLDER_MAP = load_placeholder_map()

BRAND_KEYS = {k for k, _, _, _ in BRAND_META}
CATEGORY_SHEET = "类目数据源"


def resolve_brand_key(raw_key: str | None, name_key_col: str | None = None) -> str | None:
    """Excel 行 → 标准 name_key（支持 jc_* 占位与 name_key 列）"""
    if name_key_col:
        nk = name_key_col.strip().lower()
        if nk in BRAND_KEYS:
            return nk
    if not raw_key:
        return None
    key = str(raw_key).strip().lower()
    key = PLACEHOLDER_MAP.get(key, key)
    return key if key in BRAND_KEYS else None


def parse_period(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    s = str(v).strip()
    m = re.match(r"(\d{4})年(\d{1,2})月", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    if len(s) >= 7 and s[4] == "-":
        return s[:7]
    return None


def weighted_avg(pairs: list[tuple[float, float | None]]) -> float | None:
    num = den = 0.0
    for weight, val in pairs:
        if weight and val is not None:
            num += weight * float(val)
            den += weight
    return num / den if den else None


def build_metrics_row(key: str, period: str, chs: dict) -> dict:
    jd = chs.get("JD", {})
    tm = chs.get("A-TM(大贸+国内现货)", {})
    tb = chs.get("A-TB(非国际)", {})
    dy = chs.get("N-DY(非国际)", {})
    return {
        "name_key": key,
        "period_type": "monthly",
        "period_value": period,
        "gmv": wan(jd.get("gmv_yuan")),
        "gmv_yoy": pct(jd.get("gmv_yoy")),
        "sales_volume": jd_sales_int(jd),
        "sales_volume_yoy": pct(jd.get("sales_yoy")),
        "jd_share": pct(jd.get("share")),
        "jd_share_wow": pct(jd.get("share_chg")),
        "tmall_share": pct(tm.get("share")),
        "douyin_share": pct(dy.get("share")),
        "taobao_share": pct(tb.get("share")),
        "channel_growth_jd": pct(jd.get("gmv_yoy")),
        "channel_growth_tmall": pct(tm.get("gmv_yoy")),
        "channel_growth_douyin": pct(dy.get("gmv_yoy")),
        "channel_growth_taobao": pct(tb.get("gmv_yoy")),
        "sku_count": jd.get("sku"),
    }


def rollup_brand_metrics_from_category(xlsx_path: Path, existing: set[tuple[str, str]]) -> list[dict]:
    """「类目数据源」按品牌×月×渠道汇总；补「数据源」缺失月份（如 2025-05）。"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if CATEGORY_SHEET not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[CATEGORY_SHEET]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {h: i for i, h in enumerate(header) if h}

    buckets: dict[tuple[str, str, str], dict] = defaultdict(
        lambda: {
            "gmv_yuan": 0.0,
            "sales": 0.0,
            "sku": 0,
            "share_pairs": [],
            "share_chg_pairs": [],
            "gmv_yoy_pairs": [],
            "sales_yoy_pairs": [],
        }
    )

    for r in ws.iter_rows(min_row=2, values_only=True):
        nk_col = r[idx["name_key"]] if "name_key" in idx else None
        key = resolve_brand_key(r[idx["brand"]], str(nk_col) if nk_col else None)
        if not key:
            continue
        period = parse_period(r[idx["时间"]])
        if not period or (key, period) in existing:
            continue
        ch_label = r[idx["渠道"]]
        if ch_label not in CHANNELS:
            continue
        amount = r[idx["销售额"]]
        if amount is None:
            continue
        sales_yuan = float(amount)
        b = buckets[(key, period, ch_label)]
        b["gmv_yuan"] += sales_yuan
        if r[idx["销量"]] is not None:
            b["sales"] += float(r[idx["销量"]])
        if r[idx["商品数"]] is not None:
            b["sku"] += int(float(r[idx["商品数"]]))
        if r[idx["市占率"]] is not None:
            b["share_pairs"].append((sales_yuan, float(r[idx["市占率"]])))
        if r[idx["市占率同环比"]] is not None:
            b["share_chg_pairs"].append((sales_yuan, float(r[idx["市占率同环比"]])))
        if r[idx["销售额同环比"]] is not None:
            b["gmv_yoy_pairs"].append((sales_yuan, float(r[idx["销售额同环比"]])))
        if r[idx["销量同环比"]] is not None:
            b["sales_yoy_pairs"].append((sales_yuan, float(r[idx["销量同环比"]])))

    wb.close()

    grouped: dict[tuple[str, str], dict] = defaultdict(dict)
    for (key, period, ch_label), b in buckets.items():
        grouped[(key, period)][ch_label] = {
            "gmv_yuan": b["gmv_yuan"],
            "gmv_yoy": weighted_avg(b["gmv_yoy_pairs"]),
            "sales": b["sales"] if b["sales"] else None,
            "sales_yoy": weighted_avg(b["sales_yoy_pairs"]),
            "share": weighted_avg(b["share_pairs"]),
            "share_chg": weighted_avg(b["share_chg_pairs"]),
            "sku": b["sku"] if b["sku"] else None,
        }

    return [build_metrics_row(key, period, chs) for (key, period), chs in sorted(grouped.items())]


def pct(v):
    if v is None:
        return None
    return round(float(v) * 100, 2)


def wan(yuan):
    if yuan is None:
        return None
    return round(float(yuan) / 10000, 2)


def jd_sales_int(jd):
    sales = jd.get("sales")
    if sales is None:
        return None
    return int(round(float(sales)))


def fmt(v):
    if v is None or v == "":
        return ""
    if isinstance(v, float):
        return f"{v:.2f}".rstrip("0").rstrip(".") if v % 1 else f"{int(v)}"
    return str(v)


def transform(xlsx_path: Path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header) if h}

    key_col = "brand"
    if key_col not in idx:
        key_col = "barnd  key" if "barnd  key" in idx else "brand key"
    if key_col not in idx:
        raise KeyError("数据源缺少 brand / brand key 列")
    raw = defaultdict(dict)
    brand_names = {}

    for r in rows[1:]:
        nk_col = r[idx["name_key"]] if "name_key" in idx else None
        key = resolve_brand_key(r[idx[key_col]], str(nk_col) if nk_col else None)
        if not key:
            continue
        brand_names[key] = r[idx["标准品牌"]] if "标准品牌" in idx else key
        t = r[idx["时间"]]
        period = t.strftime("%Y-%m") if isinstance(t, datetime) else str(t)[:7]
        ch = r[idx["渠道"]]
        if ch not in CHANNELS:
            continue
        raw[(key, period)][ch] = {
            "gmv_yuan": float(r[idx["销售额"]]),
            "gmv_yoy": r[idx["销售额同环比"]],
            "sales": float(r[idx["销量"]]) if r[idx["销量"]] is not None else None,
            "sales_yoy": r[idx["销量同环比"]],
            "share": r[idx["市占率"]],
            "share_chg": r[idx["市占率同环比"]],
            "sku": int(r[idx["商品数"]]) if r[idx["商品数"]] is not None else None,
        }

    out = []
    for (key, period), chs in sorted(raw.items()):
        out.append(build_metrics_row(key, period, chs))

    return out, brand_names


def write_csv(rows, path: Path):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({k: fmt(row.get(k)) for k in CSV_FIELDS})


def write_brands(path: Path):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["name", "name_key", "level", "responsible", "baseline_freq"],
        )
        w.writeheader()
        for name_key, name, level, responsible in BRAND_META:
            w.writerow(
                {
                    "name": name,
                    "name_key": name_key,
                    "level": level,
                    "responsible": responsible,
                    "baseline_freq": "季度/次",
                }
            )


def main():
    xlsx = LOCAL_XLSX if LOCAL_XLSX.exists() else DEFAULT_XLSX
    if len(sys.argv) > 1:
        xlsx = Path(sys.argv[1])
    if not xlsx.exists():
        raise SystemExit(f"找不到数据源: {xlsx}")

    rows, brand_names = transform(xlsx)
    existing = {(r["name_key"], r["period_value"]) for r in rows}
    rollup_rows = rollup_brand_metrics_from_category(xlsx, existing)
    if rollup_rows:
        rows.extend(rollup_rows)
        rows.sort(key=lambda r: (r["name_key"], r["period_value"]))
    write_csv(rows, OUT_CSV)
    write_brands(OUT_BRANDS)
    print(f"已转换 {len(rows)} 行 → {OUT_CSV}" + (f"（类目汇总补充 {len(rollup_rows)} 行）" if rollup_rows else ""))
    print(f"品牌主数据 → {OUT_BRANDS}")
    print("品牌:", ", ".join(f"{k}({brand_names.get(k, '')})" for k, _, _, _ in BRAND_META))


if __name__ == "__main__":
    main()
